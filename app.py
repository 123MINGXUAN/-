# -*- coding: utf-8 -*-
"""
退款率预测 - Web页面
启动: streamlit run app.py
"""
import streamlit as st
import pandas as pd
import numpy as np
import re, os, pickle, io
from datetime import datetime

# ==================== 配置 ====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MODEL_DIR = os.path.join(BASE_DIR, 'models')
MODEL_FILE = os.path.join(MODEL_DIR, 'refund_model.pkl')
PRODUCT_RATES_FILE = os.path.join(MODEL_DIR, 'product_refund_rates.pkl')

# 新品退款率匹配（仅使用预生成的 pkl 缓存）
PRODUCT_ANALYSIS_CACHE = os.path.join(MODEL_DIR, 'product_analysis_cache.pkl')
# ================================================


# ==================== 新品退款率匹配 ====================
@st.cache_data(ttl=3600)
def load_product_analysis():
    """加载商品分析数据库（从预生成的 pkl 缓存）"""
    if os.path.exists(PRODUCT_ANALYSIS_CACHE):
        try:
            df = pd.read_pickle(PRODUCT_ANALYSIS_CACHE)
            return df
        except Exception as e:
            print(f"缓存加载失败: {e}")
    return pd.DataFrame()


def extract_keywords(name):
    """从商品名提取关键词（领型、袖型、品类）"""
    if not isinstance(name, str):
        return set()

    keywords = set()
    # 领型
    for pattern in ['方领', 'V领', '圆领', 'Polo领', '小高领', '一字领', '斜领', '不规则']:
        if pattern in name:
            keywords.add(pattern)
    # 袖型
    for pattern in ['短袖', '中袖', '长袖', '飞袖', '背心', '露脐']:
        if pattern in name:
            keywords.add(pattern)
    # 品类
    for pattern in ['短袖', '背心', 'T恤', '卫衣', '外套']:
        if pattern in name:
            keywords.add(pattern)

    return keywords


def match_new_product(product_name, product_cost, product_db, prod_rates):
    """
    为新品匹配相似商品的退款率
    返回: (matched_rate, match_source, matched_products)
    """
    # 如果已有历史退款率，直接返回
    if product_name in prod_rates and prod_rates[product_name] > 0:
        return prod_rates[product_name], '历史数据', [product_name]

    if product_db.empty:
        return 0, '无商品库', []

    # 提取新品关键词
    new_keywords = extract_keywords(product_name)
    new_cost = product_cost if product_cost and product_cost > 0 else None

    # 计算相似度分数
    scores = []
    for idx, row in product_db.iterrows():
        score = 0
        reasons = []

        # 关键词匹配（权重最高）
        db_keywords = row['关键词'] if isinstance(row['关键词'], set) else set()
        common_keywords = new_keywords & db_keywords
        if common_keywords:
            score += len(common_keywords) * 30  # 每个匹配关键词30分
            reasons.append(f"关键词:{','.join(common_keywords)}")

        # 价格区间匹配（±30%）
        if new_cost and pd.notna(row['成本价']) and row['成本价'] > 0:
            cost_ratio = abs(new_cost - row['成本价']) / row['成本价']
            if cost_ratio <= 0.3:
                score += 20
                reasons.append(f"价格相近({row['成本价']:.1f})")

        # 有销量的商品优先
        if row['月销量'] > 10:
            score += 10
        if row['月销量'] > 50:
            score += 10

        # 有退款率数据
        if row['退款率'] > 0:
            score += 5

        if score > 0:
            scores.append((idx, score, row['退款率'], reasons))

    if not scores:
        # 没有匹配，用全店平均退款率
        avg_refund = product_db[product_db['月销量'] > 0]['退款率'].mean()
        return avg_refund if avg_refund > 0 else 0.32, '全店平均', []

    # 按分数排序，取前5个加权平均
    scores.sort(key=lambda x: x[1], reverse=True)
    top_matches = scores[:5]

    total_score = sum(s[1] for s in top_matches)
    weighted_rate = sum(s[2] * s[1] for s in top_matches) / total_score

    matched_names = [product_db.loc[s[0], '款式名/商品名'] for s in top_matches]
    match_reasons = [f"{name}({','.join(s[3])})" for name, s in zip(matched_names, top_matches)]

    return weighted_rate, f'相似匹配({len(top_matches)}款)', matched_names


# ================================================

FEATURE_COLS = [
    'sales', 'orders', 'avg_price', 'unique_products',
    'weekday', 'day', 'is_weekend', 'is_618',
    'days_since_618_end', 'post_618', 'post_618_days',
    'refund_ma3', 'refund_ma5', 'refund_ma7', 'refund_ma14',
    'orders_ma3', 'orders_ma7', 'orders_ma14',
    'sales_ma3', 'sales_ma7', 'sales_ma14',
    'refund_lag1', 'refund_lag2', 'refund_lag3', 'refund_lag5',
    'sales_lag1', 'sales_lag2', 'sales_lag3',
    'refund_momentum1', 'refund_momentum3', 'refund_accel',
    'refund_std7', 'refund_trend',
    'sales_ratio', 'refund_declining',
    'product_mix_refund',
    'wd_0', 'wd_1', 'wd_2', 'wd_3', 'wd_4', 'wd_5', 'wd_6',
]


# ==================== 自定义CSS ====================
CUSTOM_CSS = """
<style>
    /* 全局字体 */
    .stApp {
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    }

    /* 隐藏默认元素 */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}

    /* 主标题样式 */
    .main-title {
        font-size: 2.2rem;
        font-weight: 700;
        color: #1a1a2e;
        text-align: center;
        margin-bottom: 0.5rem;
    }
    .sub-title {
        font-size: 1rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
    }

    /* 卡片样式 */
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 16px;
        padding: 1.5rem 1rem;
        color: white;
        text-align: center;
        box-shadow: 0 10px 30px rgba(102, 126, 234, 0.3);
        transition: transform 0.2s;
        height: 140px;
        display: flex;
        flex-direction: column;
        justify-content: center;
        overflow: hidden;
    }
    .metric-card:hover {
        transform: translateY(-2px);
    }
    .metric-card .label {
        font-size: 0.85rem;
        opacity: 0.9;
        margin-bottom: 0.5rem;
    }
    .metric-card .value {
        font-size: 1.8rem;
        font-weight: 700;
        line-height: 1.2;
    }
    .metric-card .delta {
        font-size: 0.75rem;
        opacity: 0.8;
        margin-top: 0.3rem;
    }

    .metric-card.green {
        background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%);
        box-shadow: 0 10px 30px rgba(17, 153, 142, 0.3);
    }
    .metric-card.orange {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        box-shadow: 0 10px 30px rgba(245, 87, 108, 0.3);
    }
    .metric-card.blue {
        background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
        box-shadow: 0 10px 30px rgba(79, 172, 254, 0.3);
    }

    /* 区块标题 */
    .section-title {
        font-size: 1.2rem;
        font-weight: 600;
        color: #1a1a2e;
        padding: 0.8rem 0;
        border-bottom: 2px solid #667eea;
        margin: 1.5rem 0 1rem 0;
    }

    /* 上传区域美化 */
    .upload-area {
        background: #f8f9fa;
        border: 2px dashed #667eea;
        border-radius: 12px;
        padding: 2rem;
        text-align: center;
        margin: 1rem 0;
    }

    /* 输入框美化 */
    .stNumberInput > div > div > input {
        border-radius: 8px;
        border: 1px solid #ddd;
        padding: 0.5rem 1rem;
    }

    /* 按钮美化 */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.7rem 2rem;
        font-weight: 600;
        width: 100%;
        transition: opacity 0.2s;
    }
    .stDownloadButton > button:hover {
        opacity: 0.9;
    }

    /* 表格美化 */
    .stDataFrame {
        border-radius: 8px;
        overflow: hidden;
    }

    /* 侧边栏 */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a1a2e 0%, #16213e 100%);
    }
    section[data-testid="stSidebar"] .stMarkdown p,
    section[data-testid="stSidebar"] .stMarkdown h1,
    section[data-testid="stSidebar"] .stMarkdown h2,
    section[data-testid="stSidebar"] .stMarkdown h3 {
        color: white;
    }

    /* 分隔线 */
    hr {
        border: none;
        border-top: 1px solid #e0e0e0;
        margin: 1.5rem 0;
    }

    /* 信息提示框 */
    .info-box {
        background: linear-gradient(135deg, #667eea15 0%, #764ba215 100%);
        border-left: 4px solid #667eea;
        border-radius: 0 8px 8px 0;
        padding: 1rem 1.5rem;
        margin: 1rem 0;
    }
    .info-box p {
        margin: 0;
        color: #444;
        font-size: 0.9rem;
    }
</style>
"""


WEEKDAY_CN = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']


def fmt_date_cn(ts):
    if ts is None:
        return '未选择日期'
    return f'{ts.year}年{ts.month}月{ts.day}日 {WEEKDAY_CN[ts.weekday()]}'


@st.cache_resource
def load_model():
    with open(MODEL_FILE, 'rb') as f:
        model = pickle.load(f)
    with open(PRODUCT_RATES_FILE, 'rb') as f:
        prod_rates = pickle.load(f)
    return model, prod_rates


def load_target_file(file_obj, filename, manual_date=None):
    df = pd.read_excel(file_obj)

    target_date = None

    date_cols = ['日期', '订单日期', '发货日期', '付款日期', '日期时间']
    for c in date_cols:
        if c in df.columns:
            vals = df[c].dropna()
            if len(vals) > 0:
                try:
                    target_date = pd.to_datetime(vals.iloc[0])
                    break
                except:
                    pass

    if target_date is None and manual_date:
        target_date = manual_date

    if '数据类型' in df.columns:
        main_rows = df[df['数据类型'] == '主表'].copy()
        if len(main_rows) == 0:
            main_rows = df.copy()
    else:
        main_rows = df.copy()

    name_col = None
    for c in ['款式名称/商品名称', '商品简称', '商品名称']:
        if c in main_rows.columns:
            name_col = c
            break
    if name_col is None:
        name_col = main_rows.columns[4] if len(main_rows.columns) > 4 else main_rows.columns[0]

    product_sales = main_rows.groupby(name_col).agg(
        sales=('销售金额', 'sum'),
        quantity=('销售数量', 'sum'),
        cost=('销售成本', 'sum'),
    ).reset_index()
    product_sales.columns = ['product', 'sales', 'quantity', 'cost']

    total_sales = product_sales['sales'].sum()
    total_cost = product_sales['cost'].sum()
    total_qty = product_sales['quantity'].sum()
    n_products = len(product_sales)

    return {
        'date': target_date,
        'product_sales': product_sales,
        'total_sales': total_sales,
        'total_cost': total_cost,
        'total_quantity': total_qty,
        'unique_products': n_products,
        'raw_df': df,
    }


def add_features(daily):
    daily['weekday'] = daily['date'].dt.weekday
    daily['day'] = daily['date'].dt.day
    daily['is_weekend'] = (daily['weekday'] >= 5).astype(int)
    daily['is_618'] = ((daily['date'].dt.month == 6) & (daily['day'] <= 20)).astype(int)
    daily['days_since_618_end'] = (daily['date'] - pd.Timestamp('2026-06-21')).dt.days
    daily['post_618'] = (daily['days_since_618_end'] > 0).astype(int)
    daily['post_618_days'] = np.maximum(daily['days_since_618_end'], 0)

    for w in [3, 5, 7, 14]:
        daily[f'refund_ma{w}'] = daily['refund_rate'].rolling(w, min_periods=1).mean()
        daily[f'orders_ma{w}'] = daily['orders'].rolling(w, min_periods=1).mean()
        daily[f'sales_ma{w}'] = daily['sales'].rolling(w, min_periods=1).mean()

    for lag in [1, 2, 3, 5]:
        daily[f'refund_lag{lag}'] = daily['refund_rate'].shift(lag)
        daily[f'sales_lag{lag}'] = daily['sales'].shift(lag)

    daily['refund_momentum1'] = daily['refund_rate'] - daily['refund_rate'].shift(1)
    daily['refund_momentum3'] = daily['refund_rate'] - daily['refund_rate'].shift(3)
    daily['refund_accel'] = daily['refund_momentum1'] - daily['refund_momentum1'].shift(1)
    daily['refund_std7'] = daily['refund_rate'].rolling(7, min_periods=1).std().fillna(0)
    daily['refund_trend'] = daily['refund_ma3'] - daily['refund_ma7']
    daily['sales_ratio'] = np.where(
        daily['sales_ma7'] > 0, daily['sales'] / daily['sales_ma7'], 1
    )
    daily['refund_declining'] = (
            (daily['refund_rate'] < daily['refund_lag1']) &
            (daily['refund_lag1'] < daily['refund_lag2'])
    ).astype(int)

    for wd in range(7):
        daily[f'wd_{wd}'] = (daily['weekday'] == wd).astype(int)

    return daily.fillna(0)


def predict(target_info, model, prod_rates, product_db=None):
    ps = target_info['product_sales'].copy()

    # 新品退款率匹配
    new_products_info = []
    if product_db is not None and not product_db.empty:
        hist_refund = []
        match_sources = []
        for _, row in ps.iterrows():
            rate, source, matched = match_new_product(
                row['product'], row['cost'], product_db, prod_rates
            )
            hist_refund.append(rate)
            match_sources.append(source)
            if source != '历史数据':
                new_products_info.append({
                    'product': row['product'],
                    'matched_rate': rate,
                    'source': source,
                    'matched_products': matched,
                })
        ps['hist_refund'] = hist_refund
        ps['match_source'] = match_sources
    else:
        ps['hist_refund'] = ps['product'].map(prod_rates).fillna(0)
        ps['match_source'] = np.where(ps['hist_refund'] > 0, '历史数据', '无数据')

    ps['weighted'] = ps['hist_refund'] * ps['sales']
    total_sales = ps['sales'].sum()
    pmr = ps['weighted'].sum() / total_sales if total_sales > 0 else 0

    target_dt = target_info['date']
    date_is_none = target_dt is None
    if date_is_none:
        target_dt = pd.Timestamp.now().normalize()

    new_row = pd.DataFrame([{
        'date': target_dt,
        'orders': target_info['total_quantity'],
        'sales': target_info['total_sales'],
        'returns': 0,
        'cost': target_info['total_cost'],
        'returned_orders': 0,
        'avg_price': target_info['total_sales'] / target_info['total_quantity'] if target_info['total_quantity'] > 0 else 0,
        'unique_products': target_info['unique_products'],
        'refund_rate': np.nan,
        'product_mix_refund': pmr,
    }])

    empty_daily = pd.DataFrame(columns=['date', 'orders', 'sales', 'returns', 'cost',
                                        'returned_orders', 'avg_price', 'unique_products',
                                        'refund_rate', 'product_mix_refund'])
    daily = pd.concat([empty_daily, new_row], ignore_index=True)
    daily['date'] = pd.to_datetime(daily['date'])
    daily = add_features(daily)
    daily = daily.fillna(0)

    pred_row = daily[daily['date'] == target_dt]
    X_pred = pred_row[FEATURE_COLS]
    pred_rate = np.clip(model.predict(X_pred)[0], 0, 1)

    sales = target_info['total_sales']
    cost = target_info['total_cost']
    gross = sales - cost
    gross_rate = gross / sales if sales > 0 else 0
    refund_amt = sales * pred_rate
    net_profit = gross - refund_amt

    return {
        'date': target_dt,
        'date_is_none': date_is_none,
        'sales': sales,
        'cost': cost,
        'orders': target_info['total_quantity'],
        'gross': gross,
        'gross_rate': gross_rate,
        'pred_rate': pred_rate,
        'refund_amt': refund_amt,
        'net_profit': net_profit,
        'product_mix_refund': pmr,
        'product_sales': ps,
        'new_products_info': new_products_info,
    }


def to_excel(result):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = '店铺利润表'

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    title_font = Font(bold=True, size=14, color='FFFFFF', name='微软雅黑')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_font = Font(bold=True, size=10, color='FFFFFF', name='微软雅黑')
    data_font = Font(size=10, name='微软雅黑')
    light_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    date_str = fmt_date_cn(result['date'])

    ws.merge_cells('A1:P1')
    ws['A1'] = '店铺利润预测报表'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:P2')
    ws['A2'] = f'预测日期: {date_str} | 店铺: DY43白剑虹XIUSS（王莉）艾甜甜 | 平台: 抖音'
    ws['A2'].font = Font(size=9, color='666666', name='微软雅黑')
    ws['A2'].alignment = center_align
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 8

    headers = ['店铺类型', '平台', '店铺名称', '渠道编码', '日期',
               '销售额', '订单量', '销售成本', '毛利', '毛利率',
               '预测退款率', '预估退款金额', '运营成本', '物流成本', '广告支出', '预估净利润']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    data_row = ['自营店', '抖音', 'DY43白剑虹XIUSS', 'DY001', date_str,
                result['sales'], result['orders'], result['cost'], result['gross'], result['gross_rate'],
                result['pred_rate'], result['refund_amt'], 0, 0, 0, result['net_profit']]

    for col, value in enumerate(data_row, 1):
        cell = ws.cell(row=5, column=col, value=value)
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = white_fill
        if col in [10, 11]:
            cell.number_format = '0.00%'
        elif col in [6, 8, 9, 12, 13, 14, 15, 16]:
            cell.number_format = '#,##0.00'

    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 25
    ws.row_dimensions[6].height = 8

    ws.merge_cells('A7:P7')
    ws['A7'] = '利润分析'
    ws['A7'].font = Font(bold=True, size=11, color='1F4E79', name='微软雅黑')
    ws['A7'].alignment = Alignment(horizontal='left', vertical='center')

    summary_headers = ['项目', '金额', '占销售额比', '说明']
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    summary_data = [
        ['销售额', result['sales'], '100%', '当日总销售额'],
        ['销售成本', result['cost'],
         f'{result["cost"] / result["sales"] * 100:.1f}%' if result['sales'] > 0 else '0%', '进货成本'],
        ['毛利', result['gross'], f'{result["gross_rate"] * 100:.1f}%', '销售额 - 成本'],
        ['预估退款金额', result['refund_amt'], f'{result["pred_rate"] * 100:.1f}%', '销售额 × 预测退款率'],
        ['预估净利润', result['net_profit'],
         f'{result["net_profit"] / result["sales"] * 100:.1f}%' if result['sales'] > 0 else '0%', '毛利 - 退款'],
    ]

    for row_idx, row_data in enumerate(summary_data, 9):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = light_fill if row_idx % 2 == 0 else white_fill
            if col == 2 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    ws.row_dimensions[8].height = 25
    for r in range(9, 14):
        ws.row_dimensions[r].height = 22

    ws.row_dimensions[14].height = 8
    ws.merge_cells('A15:P15')
    ws['A15'] = '商品组合分析'
    ws['A15'].font = Font(bold=True, size=11, color='1F4E79', name='微软雅黑')
    ws['A15'].alignment = Alignment(horizontal='left', vertical='center')

    prod_headers = ['商品名称', '销售额', '销售额占比', '历史退款率', '加权退款贡献']
    for col, header in enumerate(prod_headers, 1):
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ps = result['product_sales'].sort_values('sales', ascending=False)
    for row_idx, (_, row) in enumerate(ps.iterrows(), 17):
        pct = row['sales'] / result['sales'] * 100 if result['sales'] > 0 else 0
        weighted = row['hist_refund'] * pct / 100
        values = [row['product'], row['sales'], f'{pct:.2f}%',
                  f'{row["hist_refund"]:.2%}', f'{weighted:.2%}']
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = light_fill if row_idx % 2 == 0 else white_fill
            if col == 2 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    col_widths = [12, 8, 20, 12, 12, 12, 10, 12, 12, 10, 12, 12, 10, 10, 10, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def render_metric_card(label, value, delta=None, card_type="purple"):
    """渲染一个渐变色指标卡片"""
    css_class = f"metric-card {card_type}"
    delta_html = f'<div class="delta">{delta}</div>' if delta else ''
    return f"""
    <div class="{css_class}">
        <div class="label">{label}</div>
        <div class="value">{value}</div>
        {delta_html}
    </div>
    """


# ==================== 页面 ====================
st.set_page_config(
    page_title="退款率预测",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

from auth import check_login, logout_button
check_login()
logout_button()
if 'display_name' in st.session_state:
    st.sidebar.info(f"当前用户: {st.session_state['display_name']}")

# 顶部标题
st.markdown('<div class="main-title">电商退款率预测系统</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">基于机器学习的退款率预测 & 利润分析工具</div>', unsafe_allow_html=True)

if not os.path.exists(MODEL_FILE):
    st.error("模型文件不存在: models/refund_model.pkl\n\n请联系管理员部署模型文件")
    st.stop()

if not os.path.exists(PRODUCT_RATES_FILE):
    st.error("商品退款率文件不存在: models/product_refund_rates.pkl\n\n请联系管理员部署模型文件")
    st.stop()

model, prod_rates = load_model()

# 加载商品分析数据库（用于新品匹配）
product_db = load_product_analysis()
if not product_db.empty:
    st.sidebar.success(f"已加载 {len(product_db)} 条商品数据（新品匹配已启用）")
else:
    st.sidebar.warning("商品分析缓存未找到，新品将使用全店平均退款率")

# ==================== 上传区域 ====================
st.markdown('<div class="section-title">上传数据文件</div>', unsafe_allow_html=True)

col_upload, col_gap, col_result = st.columns([1, 0.1, 2])

with col_upload:
    st.markdown("""
    <div class="info-box">
        <p><strong>支持文件：</strong>聚水潭「款和明细」报表（xlsx格式）</p>
        <p>支持同时上传多个文件批量预测</p>
        <p>日期从文件内容识别，如无日期可手动选择或不选</p>
    </div>
    """, unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "点击选择或拖拽文件到此处（支持多选）",
        type=['xlsx'],
        label_visibility="collapsed",
        accept_multiple_files=True,
    )

    st.markdown("---")
    st.markdown('<div class="section-title">预测日期（可选）</div>', unsafe_allow_html=True)
    manual_date_val = st.date_input("选择预测日期", value=None, format="YYYY-MM-DD")

    st.markdown("---")
    st.markdown('<div class="section-title">成本录入（可选）</div>', unsafe_allow_html=True)

    ops_cost = st.number_input("运营成本（元/天）", min_value=0.0, value=0.0, step=100.0, format="%.2f")
    logistics_cost = st.number_input("物流成本（元/天）", min_value=0.0, value=0.0, step=100.0, format="%.2f")
    ad_cost = st.number_input("广告支出（元/天）", min_value=0.0, value=0.0, step=100.0, format="%.2f")

# ==================== 结果展示 ====================
with col_result:
    if uploaded_files:
        all_results = []
        errors = []

        for file in uploaded_files:
            try:
                manual_ts = pd.to_datetime(manual_date_val) if manual_date_val else None
                target_info = load_target_file(file, file.name, manual_date=manual_ts)
                result = predict(target_info, model, prod_rates, product_db)

                total_cost = result['refund_amt'] + ops_cost + logistics_cost + ad_cost
                final_profit = result['gross'] - total_cost
                result['final_profit'] = final_profit
                result['ops_cost'] = ops_cost
                result['logistics_cost'] = logistics_cost
                result['ad_cost'] = ad_cost

                ps = result['product_sales'].copy()
                total_sales = result['sales']

                total_weighted = (ps['hist_refund'] * ps['sales']).sum()
                if total_weighted > 0:
                    ps['refund_share'] = result['refund_amt'] * (ps['hist_refund'] * ps['sales']) / total_weighted
                else:
                    ps['refund_share'] = result['refund_amt'] * ps['sales'] / total_sales if total_sales > 0 else 0

                ps['ops_share'] = ops_cost * ps['sales'] / total_sales if total_sales > 0 else 0
                ps['logistics_share'] = logistics_cost * ps['sales'] / total_sales if total_sales > 0 else 0
                ps['ad_share'] = ad_cost * ps['sales'] / total_sales if total_sales > 0 else 0
                ps['product_gross'] = ps['sales'] - ps['cost']
                ps['product_net'] = ps['product_gross'] - ps['refund_share'] - ps['ops_share'] - ps['logistics_share'] - ps['ad_share']
                ps['sales_ratio'] = ps['sales'] / total_sales if total_sales > 0 else 0
                ps['weighted_refund_contrib'] = ps['hist_refund'] * ps['sales_ratio']

                result['product_breakdown'] = ps
                all_results.append(result)
            except Exception as e:
                errors.append(f"{file.name}: {str(e)}")

        if errors:
            for err in errors:
                st.warning(f"跳过文件: {err}")

        if all_results:
            all_results.sort(key=lambda x: x['date'] if x['date'] is not None else pd.Timestamp.min)

            total_sales_all = sum(r['sales'] for r in all_results)
            total_refund_all = sum(r['refund_amt'] for r in all_results)
            total_profit_all = sum(r['final_profit'] for r in all_results)
            avg_rate_all = total_refund_all / total_sales_all if total_sales_all > 0 else 0
            has_date = not all(r.get('date_is_none', False) for r in all_results)

            if has_date:
                if len(all_results) == 1:
                    date_display = fmt_date_cn(all_results[0]['date'])
                else:
                    date_display = f"{fmt_date_cn(all_results[0]['date'])} ~ {fmt_date_cn(all_results[-1]['date'])}"
            else:
                date_display = '未选择日期'

            st.markdown(f'<div class="section-title">预测结果 — {date_display}</div>', unsafe_allow_html=True)

            card1, card2, card3, card4 = st.columns(4)
            with card1:
                st.markdown(render_metric_card("预测退款率", f"{avg_rate_all:.2%}", card_type="purple"), unsafe_allow_html=True)
            with card2:
                st.markdown(render_metric_card("销售额", f"¥{total_sales_all:,.0f}", card_type="blue"), unsafe_allow_html=True)
            with card3:
                st.markdown(render_metric_card("预估退款金额", f"¥{total_refund_all:,.0f}", card_type="orange"), unsafe_allow_html=True)
            with card4:
                profit_pct = f"{total_profit_all/total_sales_all*100:.1f}%" if total_sales_all > 0 else "0%"
                st.markdown(render_metric_card("预估净利润", f"¥{total_profit_all:,.0f}", delta=f"利润率 {profit_pct}", card_type="green"), unsafe_allow_html=True)

            st.markdown("---")

            for idx, r in enumerate(all_results):
                if len(all_results) > 1:
                    if r.get('date_is_none', False):
                        st.markdown(f'<div class="section-title">文件: {uploaded_files[idx].name}</div>', unsafe_allow_html=True)
                    else:
                        st.markdown(f'<div class="section-title">{fmt_date_cn(r["date"])} — {uploaded_files[idx].name}</div>', unsafe_allow_html=True)

                st.markdown('<div class="section-title">款式明细</div>', unsafe_allow_html=True)

                ps = r['product_breakdown'].sort_values('sales', ascending=False)
                product_data = []
                for _, row in ps.iterrows():
                    product_data.append({
                        '款式名称': row['product'],
                        '销售额': row['sales'],
                        '销售成本': row['cost'],
                        '毛利': row['product_gross'],
                        '历史退款率': row['hist_refund'],
                        '预估退款金额': row['refund_share'],
                        '运营成本': row['ops_share'],
                        '物流成本': row['logistics_share'],
                        '广告支出': row['ad_share'],
                        '预估净利润': row['product_net'],
                    })

                df_products = pd.DataFrame(product_data)
                st.dataframe(
                    df_products,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        '销售额': st.column_config.NumberColumn(format='¥%.2f'),
                        '销售成本': st.column_config.NumberColumn(format='¥%.2f'),
                        '毛利': st.column_config.NumberColumn(format='¥%.2f'),
                        '历史退款率': st.column_config.NumberColumn(format='%.2f%%'),
                        '预估退款金额': st.column_config.NumberColumn(format='¥%.2f'),
                        '运营成本': st.column_config.NumberColumn(format='¥%.2f'),
                        '物流成本': st.column_config.NumberColumn(format='¥%.2f'),
                        '广告支出': st.column_config.NumberColumn(format='¥%.2f'),
                        '预估净利润': st.column_config.NumberColumn(format='¥%.2f'),
                    }
                )

                # 新品匹配信息
                new_prods = r.get('new_products_info', [])
                if new_prods:
                    st.markdown('<div class="section-title">新品匹配信息</div>', unsafe_allow_html=True)
                    new_data = []
                    for np in new_prods:
                        matched_str = ', '.join(np['matched_products'][:3]) if np['matched_products'] else '-'
                        new_data.append({
                            '新品名称': np['product'],
                            '匹配退款率': f"{np['matched_rate']:.2%}",
                            '匹配来源': np['source'],
                            '参考商品': matched_str,
                        })
                    st.dataframe(pd.DataFrame(new_data), use_container_width=True, hide_index=True)

                total_ops = r['ops_cost']
                total_log = r['logistics_cost']
                total_ad = r['ad_cost']
                has_costs = total_ops > 0 or total_log > 0 or total_ad > 0

                st.markdown(f"""
                <div style="background: #f0f2f6; padding: 0.8rem 1.5rem; border-radius: 8px; margin-top: 0.5rem; font-weight: 600;">
                    合计：销售额 ¥{r['sales']:,.2f} | 退款 ¥{r['refund_amt']:,.2f} | 净利润 ¥{r['final_profit']:,.2f} | 退款率 {r['pred_rate']:.2%}{'  | 运营 ¥{:,.2f} | 物流 ¥{:,.2f} | 广告 ¥{:,.2f}'.format(total_ops, total_log, total_ad) if has_costs else ''}
                </div>
                """, unsafe_allow_html=True)

                st.markdown("---")

            st.markdown('<div class="section-title">导出报表</div>', unsafe_allow_html=True)

            excel_buf = io.BytesIO()
            with pd.ExcelWriter(excel_buf, engine='openpyxl') as writer:
                for r_idx, r in enumerate(all_results):
                    date_label = fmt_date_cn(r['date']) if not r.get('date_is_none', False) else '未选择日期'
                    if len(all_results) == 1:
                        sheet_name = '款式明细'
                    else:
                        sheet_name = date_label if not r.get('date_is_none', False) else f'文件{r_idx+1}'

                    ps = r['product_breakdown'].sort_values('sales', ascending=False)
                    export_df = pd.DataFrame({
                        '款式名称': ps['product'],
                        '销售额': ps['sales'],
                        '销售成本': ps['cost'],
                        '毛利': ps['product_gross'],
                        '历史退款率': ps['hist_refund'],
                        '数据来源': ps['match_source'],
                        '预估退款金额': ps['refund_share'],
                        '运营成本': ps['ops_share'],
                        '物流成本': ps['logistics_share'],
                        '广告支出': ps['ad_share'],
                        '预估净利润': ps['product_net'],
                    })
                    export_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

            excel_buf.seek(0)

            dl_col1, dl_col2, dl_col3 = st.columns([1, 2, 1])
            with dl_col2:
                if has_date:
                    file_name = f"预测报表_{all_results[0]['date'].strftime('%Y%m%d')}.xlsx"
                else:
                    file_name = "预测报表_未选择日期.xlsx"
                st.download_button(
                    label="下载 Excel 报表",
                    data=excel_buf,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

    else:
        st.markdown("""
        <div style="text-align: center; padding: 4rem 2rem; background: #f8f9fa; border-radius: 16px; margin-top: 2rem;">
            <div style="font-size: 4rem; margin-bottom: 1rem;">📊</div>
            <div style="font-size: 1.5rem; font-weight: 600; color: #1a1a2e; margin-bottom: 0.5rem;">
                请上传数据文件
            </div>
            <div style="color: #666; font-size: 0.95rem; max-width: 500px; margin: 0 auto; line-height: 1.8;">
                <p>从聚水潭导出「款和明细」报表</p>
                <p>支持同时上传多个文件批量预测</p>
                <p>自动预测退款率并计算净利润</p>
            </div>
        </div>
        """, unsafe_allow_html=True)
