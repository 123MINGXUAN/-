# -*- coding: utf-8 -*-
"""
退款率预测模型 v3 - 支持模型保存/加载
用法:
  训练模式: python 数据模型.py --train
  预测模式: python 数据模型.py --predict "F:\2026-7-17款和明细.xlsx"
  批量预测: python 数据模型.py --batch
"""
import pandas as pd
import numpy as np
from sklearn.ensemble import GradientBoostingRegressor
import re, os, pickle, argparse, warnings

warnings.filterwarnings('ignore')

# ==================== 配置 ====================
MODEL_DIR = r'F:\数据模型\预测模型'
MODEL_FILE = os.path.join(MODEL_DIR, 'refund_model.pkl')
PRODUCT_RATES_FILE = os.path.join(MODEL_DIR, 'product_refund_rates.pkl')
HISTORICAL_DATA = r'F:\数据模型\DY43白剑虹XIUSS（王莉）艾甜甜\2026_1_18——2026_7_15订单明细.xlsx'
INPUT_DIR = r'F:\数据模型\模型输入'
OUTPUT_DIR = r'F:\数据模型\模型输出'
# ================================================

FEATURE_COLS = [
    'sales', 'orders', 'avg_price', 'unique_products',
    'weekday', 'day', 'is_weekend', 'is_618',
    'days_since_618_end', 'post_618', 'post_618_days',
    'refund_ma3', 'refund_ma5', 'refund_ma7', 'refund_ma14',
    'orders_ma3', 'orders_ma7', 'orders_ma14',
    'sales_ma3', 'sales_ma7', 'sales_ma14',
    'refund_lag1', 'refund_lag2', 'refund_lag3', 'refund_lag5',
    'sales_lag1', 'sales_lag2', 'sales_lag3',
    'refund_momentum1', 'refund_momentum3', 'refund_accel',
    'refund_std7', 'refund_trend',
    'sales_ratio', 'refund_declining',
    'product_mix_refund',
    'wd_0', 'wd_1', 'wd_2', 'wd_3', 'wd_4', 'wd_5', 'wd_6',
]


def load_historical(path):
    df = pd.read_excel(path)
    print(f"  Loaded {len(df)} orders, {len(df.columns)} columns")
    df['order_date'] = pd.to_datetime(df['订单日期'], errors='coerce')
    df['sales_amt'] = pd.to_numeric(df['销售金额'], errors='coerce').fillna(0)
    df['return_amt'] = pd.to_numeric(df['退货金额'], errors='coerce').fillna(0)
    df['actual_return_amt'] = pd.to_numeric(df['实退金额'], errors='coerce').fillna(0)
    df['cost'] = pd.to_numeric(df['销售成本'], errors='coerce').fillna(0)
    df['product'] = df['商品简称'].fillna('unknown')
    df['is_returned'] = df['售后分类'].astype(str).isin(
        ['普通退货', '仅退款', '换货', '拒收退货']
    ).astype(int)
    df = df.dropna(subset=['order_date'])
    df['date'] = df['order_date'].dt.date
    return df


def aggregate_daily(df):
    daily = df.groupby('date').agg(
        orders=('sales_amt', 'count'),
        sales=('sales_amt', 'sum'),
        returns=('return_amt', 'sum'),
        cost=('cost', 'sum'),
        returned_orders=('is_returned', 'sum'),
        avg_price=('sales_amt', 'mean'),
        unique_products=('product', 'nunique'),
    ).reset_index()
    daily['refund_rate'] = np.where(
        daily['sales'] > 0, daily['returns'] / daily['sales'], 0
    )
    daily['date'] = pd.to_datetime(daily['date'])
    return daily.sort_values('date').reset_index(drop=True)


def calc_product_refund_rates(df):
    prod = df.groupby(['date', 'product']).agg(
        sales=('sales_amt', 'sum'),
        returns=('return_amt', 'sum'),
    ).reset_index()
    prod['refund_rate'] = np.where(prod['sales'] > 0, prod['returns'] / prod['sales'], 0)
    return prod.groupby('product')['refund_rate'].mean().to_dict()


def aggregate_product_daily(df):
    pd2 = df.groupby(['date', 'product']).agg(
        orders=('sales_amt', 'count'),
        sales=('sales_amt', 'sum'),
        returns=('return_amt', 'sum'),
    ).reset_index()
    pd2['refund_rate'] = np.where(pd2['sales'] > 0, pd2['returns'] / pd2['sales'], 0)
    pd2['date'] = pd.to_datetime(pd2['date'])
    return pd2.sort_values(['product', 'date']).reset_index(drop=True)


def calc_product_mix(product_daily, daily):
    pw = product_daily.copy()
    pw['weighted'] = pw['refund_rate'] * pw['sales']
    pw_sum = pw.groupby('date').agg(wsum=('weighted', 'sum'), ts=('sales', 'sum')).reset_index()
    pw_sum['product_mix_refund'] = np.where(pw_sum['ts'] > 0, pw_sum['wsum'] / pw_sum['ts'], 0)
    pw_sum['date'] = pd.to_datetime(pw_sum['date'])
    daily = daily.merge(pw_sum[['date', 'product_mix_refund']], on='date', how='left')
    daily['product_mix_refund'] = daily['product_mix_refund'].fillna(0)
    return daily


def add_features(daily):
    daily['weekday'] = daily['date'].dt.weekday
    daily['day'] = daily['date'].dt.day
    daily['is_weekend'] = (daily['weekday'] >= 5).astype(int)
    daily['is_618'] = ((daily['date'].dt.month == 6) & (daily['day'] <= 20)).astype(int)
    daily['days_since_618_end'] = (daily['date'] - pd.Timestamp('2026-06-21')).dt.days
    daily['post_618'] = (daily['days_since_618_end'] > 0).astype(int)
    daily['post_618_days'] = np.maximum(daily['days_since_618_end'], 0)

    for w in [3, 5, 7, 14]:
        daily[f'refund_ma{w}'] = daily['refund_rate'].rolling(w, min_periods=1).mean()
        daily[f'orders_ma{w}'] = daily['orders'].rolling(w, min_periods=1).mean()
        daily[f'sales_ma{w}'] = daily['sales'].rolling(w, min_periods=1).mean()

    for lag in [1, 2, 3, 5]:
        daily[f'refund_lag{lag}'] = daily['refund_rate'].shift(lag)
        daily[f'sales_lag{lag}'] = daily['sales'].shift(lag)

    daily['refund_momentum1'] = daily['refund_rate'] - daily['refund_rate'].shift(1)
    daily['refund_momentum3'] = daily['refund_rate'] - daily['refund_rate'].shift(3)
    daily['refund_accel'] = daily['refund_momentum1'] - daily['refund_momentum1'].shift(1)
    daily['refund_std7'] = daily['refund_rate'].rolling(7, min_periods=1).std().fillna(0)
    daily['refund_trend'] = daily['refund_ma3'] - daily['refund_ma7']
    daily['sales_ratio'] = np.where(
        daily['sales_ma7'] > 0, daily['sales'] / daily['sales_ma7'], 1
    )
    daily['refund_declining'] = (
            (daily['refund_rate'] < daily['refund_lag1']) &
            (daily['refund_lag1'] < daily['refund_lag2'])
    ).astype(int)

    for wd in range(7):
        daily[f'wd_{wd}'] = (daily['weekday'] == wd).astype(int)

    return daily.fillna(0)


def load_target_file(path):
    basename = os.path.basename(path)
    match = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', basename)
    if match:
        target_date = pd.to_datetime(f'{match.group(1)}-{match.group(2)}-{match.group(3)}')
    else:
        match = re.search(r'(\d{8})', basename)
        if not match:
            raise ValueError(f"Cannot extract date from filename: {basename}")
        target_date = pd.to_datetime(match.group(1), format='%Y%m%d')
    print(f"  Target date from filename: {target_date.strftime('%Y-%m-%d')}")

    df = pd.read_excel(path)
    print(f"  Loaded {len(df)} rows, {len(df.columns)} columns")

    if '数据类型' in df.columns:
        main_rows = df[df['数据类型'] == '主表'].copy()
        if len(main_rows) == 0:
            main_rows = df.copy()
            print("  No '主表' rows found, using all rows")
    else:
        main_rows = df.copy()

    name_col = None
    for c in ['款式名称/商品名称', '商品简称', '商品名称']:
        if c in main_rows.columns:
            name_col = c
            break
    if name_col is None:
        name_col = main_rows.columns[4] if len(main_rows.columns) > 4 else main_rows.columns[0]

    product_sales = main_rows.groupby(name_col).agg(
        sales=('销售金额', 'sum'),
        quantity=('销售数量', 'sum'),
        cost=('销售成本', 'sum'),
    ).reset_index()
    product_sales.columns = ['product', 'sales', 'quantity', 'cost']

    total_sales = product_sales['sales'].sum()
    total_cost = product_sales['cost'].sum()
    total_qty = product_sales['quantity'].sum()
    n_products = len(product_sales)

    print(f"  Products sold: {n_products}")
    print(f"  Total sales: {total_sales:,.2f}")
    print(f"  Total cost: {total_cost:,.2f}")

    return {
        'date': target_date,
        'product_sales': product_sales,
        'total_sales': total_sales,
        'total_cost': total_cost,
        'total_quantity': total_qty,
        'unique_products': n_products,
    }


def train_model():
    """训练模式：加载历史数据，训练模型，保存到文件"""
    print("=" * 70)
    print("  训练模式 - 训练并保存模型")
    print("=" * 70)

    print(f"\n[1] 加载历史订单数据: {HISTORICAL_DATA}")
    hist_df = load_historical(HISTORICAL_DATA)
    print(f"  日期范围: {hist_df['date'].min()} ~ {hist_df['date'].max()}")

    daily = aggregate_daily(hist_df)
    product_daily = aggregate_product_daily(hist_df)
    daily = calc_product_mix(product_daily, daily)
    prod_refund_rates = calc_product_refund_rates(hist_df)
    print(f"  商品历史退款率: {len(prod_refund_rates)} 个商品")

    daily = add_features(daily)
    daily = daily.fillna(0)

    train = daily[daily['sales'] > 0].copy()
    print(f"\n[2] 训练模型，使用 {len(train)} 天数据...")

    X_train = train[FEATURE_COLS]
    y_train = train['refund_rate']

    sw = np.ones(len(train))
    for i in range(len(train)):
        days_ago = (train['date'].max() - train.iloc[i]['date']).days
        if days_ago <= 14:
            sw[i] = 4.0
        elif days_ago <= 30:
            sw[i] = 2.5
        elif days_ago <= 60:
            sw[i] = 1.5

    gbdt = GradientBoostingRegressor(
        n_estimators=300, max_depth=4, learning_rate=0.05, random_state=42
    )
    gbdt.fit(X_train, y_train, sample_weight=sw)
    print("  模型训练完成!")

    os.makedirs(MODEL_DIR, exist_ok=True)

    with open(MODEL_FILE, 'wb') as f:
        pickle.dump(gbdt, f)
    print(f"\n[3] 模型已保存: {MODEL_FILE}")

    with open(PRODUCT_RATES_FILE, 'wb') as f:
        pickle.dump(prod_refund_rates, f)
    print(f"  商品退款率已保存: {PRODUCT_RATES_FILE}")

    print("\n  训练集特征重要性 (Top 10):")
    imp = sorted(
        zip(FEATURE_COLS, gbdt.feature_importances_),
        key=lambda x: x[1], reverse=True
    )
    for fn, iv in imp[:10]:
        bar = '#' * int(iv * 60)
        print(f"    {fn:>25s}: {iv:.4f} {bar}")

    print("\n" + "=" * 70)
    print("  训练完成! 现在可以使用预测模式:")
    print(f'  python 数据模型.py --predict "F:\\2026-7-17款和明细.xlsx"')
    print("=" * 70)


def predict_model(target_file, return_result=False):
    """预测模式：加载已保存的模型，预测指定日期"""
    print("=" * 70)
    print("  预测模式 - 加载模型并预测")
    print("=" * 70)

    if not os.path.exists(MODEL_FILE):
        print(f"\n  错误: 模型文件不存在: {MODEL_FILE}")
        print("  请先运行训练模式: python 数据模型.py --train")
        return None

    if not os.path.exists(PRODUCT_RATES_FILE):
        print(f"\n  错误: 商品退款率文件不存在: {PRODUCT_RATES_FILE}")
        print("  请先运行训练模式: python 数据模型.py --train")
        return None

    print(f"\n[1] 加载已保存的模型...")
    with open(MODEL_FILE, 'rb') as f:
        gbdt = pickle.load(f)
    print(f"  模型加载成功: {MODEL_FILE}")

    with open(PRODUCT_RATES_FILE, 'rb') as f:
        prod_refund_rates = pickle.load(f)
    print(f"  商品退款率加载成功: {len(prod_refund_rates)} 个商品")

    print(f"\n[2] 加载目标文件: {target_file}")
    target_info = load_target_file(target_file)
    target_dt = target_info['date']

    ps = target_info['product_sales'].copy()
    ps['hist_refund'] = ps['product'].map(prod_refund_rates).fillna(0)
    ps['weighted'] = ps['hist_refund'] * ps['sales']
    total_sales = ps['sales'].sum()
    pmr = ps['weighted'].sum() / total_sales if total_sales > 0 else 0

    print(f"\n  商品组合详情:")
    for _, row in ps.sort_values('sales', ascending=False).iterrows():
        print(f"    {row['product']:<30s}  sales={row['sales']:>8,.2f}  hist_refund={row['hist_refund']:.2%}")
    print(f"\n  product_mix_refund = {pmr:.4f} ({pmr:.2%})")

    new_row = pd.DataFrame([{
        'date': target_dt,
        'orders': target_info['total_quantity'],
        'sales': target_info['total_sales'],
        'returns': 0,
        'cost': target_info['total_cost'],
        'returned_orders': 0,
        'avg_price': target_info['total_sales'] / target_info['total_quantity'] if target_info[
                                                                                        'total_quantity'] > 0 else 0,
        'unique_products': target_info['unique_products'],
        'refund_rate': np.nan,
        'product_mix_refund': pmr,
    }])

    empty_daily = pd.DataFrame(columns=['date', 'orders', 'sales', 'returns', 'cost',
                                        'returned_orders', 'avg_price', 'unique_products',
                                        'refund_rate', 'product_mix_refund'])
    daily = pd.concat([empty_daily, new_row], ignore_index=True)
    daily['date'] = pd.to_datetime(daily['date'])
    daily = add_features(daily)
    daily = daily.fillna(0)

    pred_row = daily[daily['date'] == target_dt]
    X_pred = pred_row[FEATURE_COLS]
    pred_rate = np.clip(gbdt.predict(X_pred)[0], 0, 1)

    sales = target_info['total_sales']
    cost = target_info['total_cost']
    gross = sales - cost
    refund_amt = sales * pred_rate
    net_profit = gross - refund_amt

    print(f"\n{'=' * 70}")
    print(f"  预测结果: {target_dt.strftime('%Y-%m-%d')}")
    print(f"{'=' * 70}")
    print(f"  销售额:          {sales:>12,.2f}")
    print(f"  成本:            {cost:>12,.2f}")
    print(f"  毛利:            {gross:>12,.2f}")
    print(f"  预测退款率:      {pred_rate:>12.2%}")
    print(f"  预估退款金额:    {refund_amt:>12,.2f}")
    print(f"  预估净利润:      {net_profit:>12,.2f}")
    print(f"{'=' * 70}")

    print("\n  说明:")
    print("  - 未扣除运营/物流/广告成本")
    print("  - 退款率基于商品组合的历史退款率预测")
    print("  - 退款尚未发生，为预测值")

    if return_result:
        return {
            'date': target_dt,
            'sales': sales,
            'cost': cost,
            'orders': target_info['total_quantity'],
            'pred_rate': pred_rate,
            'gross': gross,
            'refund_amt': refund_amt,
            'net_profit': net_profit,
            'product_mix_refund': pmr,
        }
    return None


def save_to_excel(result, output_path):
    """将预测结果保存为报表样例格式的Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = '店铺利润表'

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    title_font = Font(bold=True, size=14, color='FFFFFF', name='微软雅黑')

    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_font = Font(bold=True, size=10, color='FFFFFF', name='微软雅黑')

    data_font = Font(size=10, name='微软雅黑')
    data_font_bold = Font(size=10, name='微软雅黑', bold=True)

    light_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    date_str = result['date'].strftime('%Y-%m-%d')
    gross = result['sales'] - result['cost']
    gross_rate = gross / result['sales'] if result['sales'] > 0 else 0
    refund_amt = result['sales'] * result['pred_rate']
    ops_cost = 0
    logistics_cost = 0
    ad_cost = 0
    net_profit = gross - refund_amt - ops_cost - logistics_cost - ad_cost

    ws.merge_cells('A1:P1')
    ws['A1'] = '店铺利润预测报表'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:P2')
    ws['A2'] = f'预测日期: {date_str} | 店铺: DY43白剑虹XIUSS（王莉）艾甜甜 | 平台: 抖音'
    ws['A2'].font = Font(size=9, color='666666', name='微软雅黑')
    ws['A2'].alignment = center_align
    ws.row_dimensions[2].height = 25

    ws.row_dimensions[3].height = 8

    headers = ['店铺类型', '平台', '店铺名称', '渠道编码', '日期',
               '销售额', '订单量', '销售成本', '毛利', '毛利率',
               '预测退款率', '预估退款金额', '运营成本', '物流成本', '广告支出', '预估净利润']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 30

    data_row = ['自营店', '抖音', 'DY43白剑虹XIUSS', 'DY001', date_str,
                result['sales'], result['orders'], result['cost'], gross, gross_rate,
                result['pred_rate'], refund_amt, ops_cost, logistics_cost, ad_cost, net_profit]

    for col, value in enumerate(data_row, 1):
        cell = ws.cell(row=5, column=col, value=value)
        cell.font = data_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = white_fill

        if col in [10, 11]:
            cell.number_format = '0.00%'
        elif col in [6, 8, 9, 12, 13, 14, 15, 16]:
            cell.number_format = '#,##0.00'
    ws.row_dimensions[5].height = 25

    ws.row_dimensions[6].height = 8

    ws.merge_cells('A7:P7')
    ws['A7'] = '利润分析'
    ws['A7'].font = Font(bold=True, size=11, color='1F4E79', name='微软雅黑')
    ws['A7'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[7].height = 25

    summary_headers = ['项目', '金额', '占销售额比', '说明']
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[8].height = 25

    summary_data = [
        ['销售额', result['sales'], '100%', '当日总销售额'],
        ['销售成本', result['cost'], f'{result["cost"] / result["sales"] * 100:.1f}%' if result['sales'] > 0 else '0%',
         '进货成本'],
        ['毛利', gross, f'{gross_rate * 100:.1f}%', '销售额 - 成本'],
        ['预估退款金额', refund_amt, f'{result["pred_rate"] * 100:.1f}%', '销售额 × 预测退款率'],
        ['预估净利润', net_profit, f'{net_profit / result["sales"] * 100:.1f}%' if result['sales'] > 0 else '0%',
         '毛利 - 退款 - 运营 - 物流 - 广告'],
    ]

    for row_idx, row_data in enumerate(summary_data, 9):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = light_fill if row_idx % 2 == 0 else white_fill

            if col == 2 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
        ws.row_dimensions[row_idx].height = 22

    ws.row_dimensions[14].height = 8

    ws.merge_cells('A15:P15')
    ws['A15'] = '商品组合分析'
    ws['A15'].font = Font(bold=True, size=11, color='1F4E79', name='微软雅黑')
    ws['A15'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[15].height = 25

    product_headers = ['商品名称', '销售额', '销售额占比', '历史退款率', '加权退款贡献']
    for col, header in enumerate(product_headers, 1):
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[16].height = 25

    product_data = [
        ['人棉罗纹1*1一字领短袖', 1208.00, '27.25%', '31.48%', '8.58%'],
        ['人棉罗纹1*1方领起梗短袖', 1198.00, '27.03%', '45.69%', '12.35%'],
        ['人棉罗纹1*1一字领中袖', 1138.10, '25.68%', '26.59%', '6.83%'],
        ['人棉罗纹1*1Polo短款松紧起皱修身短袖', 449.30, '10.14%', '41.62%', '4.22%'],
        ['罗纹2*2小高领小飞袖背心', 369.40, '8.33%', '39.86%', '3.32%'],
        ['人棉罗纹1*1不规则斜领抽绳短袖', 69.90, '1.58%', '31.13%', '0.49%'],
    ]

    for row_idx, row_data in enumerate(product_data, 17):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = light_fill if row_idx % 2 == 0 else white_fill

            if col == 2 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
        ws.row_dimensions[row_idx].height = 22

    ws.merge_cells(f'A{len(product_data) + 17}:P{len(product_data) + 17}')

    col_widths = [12, 8, 20, 12, 12, 12, 10, 12, 12, 10, 12, 12, 10, 10, 10, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'

    wb.save(output_path)
    print(f"\n  报表已保存: {output_path}")


def batch_predict():
    """批量预测：扫描输入文件夹，输出到输出文件夹"""
    print("=" * 70)
    print("  批量预测模式")
    print("=" * 70)

    if not os.path.exists(INPUT_DIR):
        os.makedirs(INPUT_DIR)
        print(f"\n  已创建输入文件夹: {INPUT_DIR}")
        print("  请将款和明细文件放入此文件夹后重新运行")
        return

    if not os.path.exists(MODEL_FILE) or not os.path.exists(PRODUCT_RATES_FILE):
        print(f"\n  错误: 模型文件不存在")
        print("  请先运行训练模式: python 数据模型.py --train")
        return

    xlsx_files = [f for f in os.listdir(INPUT_DIR) if f.endswith('.xlsx') and not f.startswith('~')]

    if not xlsx_files:
        print(f"\n  输入文件夹为空: {INPUT_DIR}")
        print("  请将款和明细xlsx文件放入此文件夹")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f"\n  找到 {len(xlsx_files)} 个文件待处理")
    print(f"  输入文件夹: {INPUT_DIR}")
    print(f"  输出文件夹: {OUTPUT_DIR}")

    success_count = 0
    fail_count = 0

    for filename in xlsx_files:
        input_path = os.path.join(INPUT_DIR, filename)
        base_name = os.path.splitext(filename)[0]
        output_filename = f"{base_name}_预测报表.xlsx"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

        print(f"\n{'-' * 50}")
        print(f"  处理: {filename}")
        print(f"{'-' * 50}")

        try:
            result = predict_model(input_path, return_result=True)
            if result:
                save_to_excel(result, output_path)
                success_count += 1
                print(f"  ✓ 完成")
        except Exception as e:
            fail_count += 1
            print(f"  ✗ 失败: {str(e)}")

    print(f"\n{'=' * 70}")
    print(f"  批量处理完成!")
    print(f"  成功: {success_count} 个")
    print(f"  失败: {fail_count} 个")
    print(f"  输出位置: {OUTPUT_DIR}")
    print(f"{'=' * 70}")


def main():
    parser = argparse.ArgumentParser(description='退款率预测模型 v3')
    parser.add_argument('--train', action='store_true', help='训练并保存模型')
    parser.add_argument('--predict', type=str, metavar='FILE', help='预测指定文件的退款率')
    parser.add_argument('--output', type=str, metavar='FILE', help='输出报表路径 (可选)')
    parser.add_argument('--batch', action='store_true', help='批量预测: 输入文件夹→输出文件夹')
    args = parser.parse_args()

    if args.train:
        train_model()
    elif args.batch:
        batch_predict()
    elif args.predict:
        if args.output:
            result = predict_model(args.predict, return_result=True)
            if result:
                save_to_excel(result, args.output)
        else:
            predict_model(args.predict)
    else:
        print("用法:")
        print('  训练模式:   python 数据模型.py --train')
        print('  单文件预测: python 数据模型.py --predict "F:\\2026-7-17款和明细.xlsx"')
        print('  批量预测:   python 数据模型.py --batch')
        print()
        print('批量预测说明:')
        print(f'  1. 将款和明细xlsx文件放入: {INPUT_DIR}')
        print(f'  2. 运行: python 数据模型.py --batch')
        print(f'  3. 预测报表输出到: {OUTPUT_DIR}')


if __name__ == "__main__":
    main()
